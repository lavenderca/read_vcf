#!/usr/bin/env python

import re
import sys
import math
import decimal
import argparse
import openpyxl

from openpyxl.styles import Font, PatternFill, Alignment

POSITION_FIELDS = [
    'Chromosome',
    'Position',
    'dbSNP ID',
    'Reference allele',
    'Alternative alleles',
    'QD',
    'Filter',
]

ANNOTATION_FIELDS = [
    'Allele',
    'Annotation',
    'Putative impact',
    'Gene Name',
    'Gene ID',
    'Feature Type',
    'Feature ID',
    'Transcript Biotype',
    'Exon or Intron Rank / Total',
    'HGVS DNA',
    'HGVS Protein',
    'cDNA Position / cDNA Length',
    'CDS Position / CDS Length',
    'Protein Position / Protein Length',
    'Distance to Feature',
    'Additional Reports?',
]

ALLELE_FREQUENCY_FIELDS = [
    'Call Rate',
    'Number of Distinct Alleles',
    'Minor Allele',
    'Major Allele',
    'Minor Allele Frequency',
    'Major Allele Frequency',
    'Fisher\'s HWE P',
    'Genotype DD Count',
    'Genotype Dd Count',
    'Genotype dd or Multi-Allelic Count',
    'Missing Genotype Count',
    'Minor Allele D Count',
    'Major Allele d Count',
    'Missing Allele Count',
]

DBNSFP_IMPACT_INFO_DICT = {
    'dbNSFP_GERP___RS': 'GERP++ RS',
    'dbNSFP_GERP___NR': 'GERP++ NR',
    'dbNSFP_MetaSVM_pred': 'MetaSVM',
    'dbNSFP_Interpro_domain': 'Interpro Domain',
    'dbNSFP_FATHMM_pred': 'FATHMM',
    'dbNSFP_LRT_pred': 'LRT',
    'dbNSFP_PROVEAN_pred': 'PROVEAN',
    'dbNSFP_phastCons100way_vertebrate': 'phastCons100way',
    'dbNSFP_CADD_phred': 'CADD',
    'dbNSFP_Polyphen2_HDIV_pred': 'Polyphen2_HDIV',
    'dbNSFP_MutationAssessor_pred': 'MutationAssessor',
    'dbNSFP_Polyphen2_HVAR_pred': 'Polyphen2_HVAR',
    'dbNSFP_SIFT_pred': 'SIFT',
}
DBNSFP_IMPACT_FIELDS = sorted(DBNSFP_IMPACT_INFO_DICT.values())

EXAC_FREQUENCY_INFO_DICT = {
    'dbNSFP_ExAC_NFE_AF': 'ExAc Non-Finnish European Frequency',
    'dbNSFP_ExAC_SAS_AF': 'ExAc South Asian Frequency',
    'dbNSFP_ExAC_Adj_AC': 'ExAc Adj. Allele Count',
    'dbNSFP_ExAC_Adj_AF': 'ExAc Adj. Allele Frequency',
    'dbNSFP_ExAC_SAS_AC': 'ExAc South Asian Count',
    'dbNSFP_ExAC_AFR_AF': 'ExAc African/African American Frequency',
    'dbNSFP_ExAC_AFR_AC': 'ExAc African/African American Count',
    'dbNSFP_ExAC_AF': 'ExAc Allele Frequency',
    'dbNSFP_ExAC_AC': 'ExAc Allele Count',
    'dbNSFP_ExAC_FIN_AC': 'ExAc Finnish Count',
    'dbNSFP_ExAC_FIN_AF': 'ExAc Finnish Frequency',
    'dbNSFP_ExAC_AMR_AF': 'ExAc American Frequency',
    'dbNSFP_ExAC_AMR_AC': 'ExAc American Count',
    'dbNSFP_ExAC_NFE_AC': 'ExAc Non-Finnish European Count',
    'dbNSFP_ExAC_EAS_AC': 'ExAc East Asian Count',
    'dbNSFP_ExAC_EAS_AF': 'ExAc East Asian Frequency',
}
EXAC_FREQUENCY_FIELDS = [
    'ExAc Allele Count',
    'ExAc Allele Frequency',
    'ExAc Adj. Allele Count',
    'ExAc Adj. Allele Frequency',
    'ExAc African/African American Count',
    'ExAc African/African American Frequency',
    'ExAc American Count',
    'ExAc American Frequency',
    'ExAc East Asian Count',
    'ExAc East Asian Frequency',
    'ExAc Finnish Count',
    'ExAc Finnish Frequency',
    'ExAc Non-Finnish European Count',
    'ExAc Non-Finnish European Frequency',
    'ExAc South Asian Count',
    'ExAc South Asian Frequency',
]
SNPEFF_IMPACT_PRIORITY = ['HIGH', 'MODERATE', 'LOW', 'MODIFIER']

COLORS = []
for fill, font in [
        ('b5854e', '000000'),
        ('9abb7d', '000000'),
        ('43301f', 'ffffff'),
        ('a7d9cf', '000000'),
        ('d3ae6f', '000000'),
    ]:
    COLORS.append({
        'fill': fill,
        'font': font,
    })


def write_headers(ws, headers):
    col_index = 1
    for i, header_group in enumerate(headers):
        fill = PatternFill(
            start_color=COLORS[i % len(COLORS)]['fill'],
            end_color=COLORS[i % len(COLORS)]['fill'],
            fill_type='solid',
        )
        font = Font(
            bold=True,
            color=COLORS[i % len(COLORS)]['font'],
        )

        cell = ws.cell(row=1, column=col_index, value=header_group['group_name'])
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(
            horizontal='center'
        )

        start_column = col_index
        for field in header_group['fields']:
            cell = ws.cell(row=2, column=col_index, value=field)
            cell.fill = fill
            cell.font = font

            col_index += 1

        ws.merge_cells(start_row=1, start_column=start_column,
                       end_row=1, end_column=(col_index - 1))


def write_fields(ws, row, fields_list, fields_dict, ref_allele=None):
    het_fill = PatternFill(
        start_color='FFEB9C',
        end_color='FFEB9C',
        fill_type='solid',
    )
    het_font = Font(
        color='9C6500',
    )

    hom_fill = PatternFill(
        start_color='FFC7CE',
        end_color='FFC7CE',
        fill_type='solid',
    )
    hom_font = Font(
        color='9C0006',
    )

    col_index = 1
    for field_group in fields_list:
        group_name = field_group['group_name']

        for field in field_group['fields']:
            key = (group_name, field)

            if key not in fields_dict:
                val = '#N/A'
            else:
                if type(fields_dict[key]) == bool:
                    val = str(fields_dict[key])
                elif fields_dict[key] == 0:
                    val = str(fields_dict[key])
                elif not fields_dict[key]:
                    val = '#N/A'
                else:
                    val = fields_dict[key]

            if type(val) == str:
                if val.isdigit():
                    val = int(val)
                else:
                    try:
                        val = float(val)
                    except ValueError:
                        pass

            cell = ws.cell(column=col_index, row=row, value=val)

            if group_name == 'Genotypes' and val != '#N/A':
                alleles = val.split('_')
                ref_count = alleles.count(ref_allele)

                if len(alleles) == ref_count:
                    pass
                elif len(alleles) == ref_count + 1:
                    cell.fill = het_fill
                    cell.font = het_font
                else:
                    cell.fill = hom_fill
                    cell.font = hom_font

            col_index += 1


def get_genotype(info_field, vcf_format, ref_allele, alt_alleles):
    genotype_index = vcf_format.strip().split(':').index('GT')
    genotype_field = info_field.strip().split(':')[genotype_index]

    allele_indices = re.split('[/\|]', genotype_field)
    possible_alleles = [ref_allele] + alt_alleles.split(',')
    genotype = []
    for i in allele_indices:
        if i == '.':
            genotype.append('?')
        else:
            genotype.append(possible_alleles[int(i)])

    return genotype


def fisher_hwe(n, Dd, D, d):
    fct = math.factorial
    Decimal = decimal.Decimal

    def calc(x):
        return (Decimal(fct(n) * fct(D) * fct(d)) / fct(2 * n)) * \
            (Decimal(2 ** x) / (fct(Decimal(D - x) / 2) * fct(x) * fct(n - Decimal(D + x) / 2)))

    n_prob = calc(Dd)
    total_prob = n_prob

    if Dd % 2 == 0:
        x_vals = range(0, Dd, 2)
    else:
        x_vals = range(1, Dd, 2)

    for x in x_vals:
        p = calc(x)
        if p <= n_prob:
            total_prob += p

    return float(total_prob)


def get_allele_frequency_values(genotypes, ref_alleles, alt_alleles):

    total_genotype_count = len(genotypes)
    total_allele_count = 0

    alleles = dict()
    for g in genotypes:
        for a in g:
            if a not in alleles:
                alleles[a] = 0
            alleles[a] += 1
            total_allele_count += 1

    if '?' in alleles:
        missing_allele_count = alleles.pop('?')
    else:
        missing_allele_count = 0

    allele_num = len(alleles)
    if allele_num > 1:
        maj_allele, min_allele = sorted(alleles.items(), key = lambda x: -x[1])[:2]
    else:
        maj_allele = alleles.items()[0]
        min_allele = ('#N/A', '#N/A')

    maj_allele_id, maj_allele_count = maj_allele
    min_allele_id, min_allele_count = min_allele
    maj_allele_freq = float(maj_allele_count) / total_allele_count
    if min_allele_id == '#N/A':
        min_allele_freq = '#N/A'
    else:
        min_allele_freq = float(min_allele_count) / total_allele_count

    genotype_DD_count = 0
    genotype_Dd_count = 0
    genotype_dd_count = 0
    missing_genotype_count = 0
    for g in genotypes:
        if '?' in g:
            missing_genotype_count += 1
        else:
            if len(g) == 2:
                if g == [min_allele_id, min_allele_id]:
                    genotype_DD_count += 1
                elif min_allele_id in g:
                    genotype_Dd_count += 1
                else:
                    genotype_dd_count += 1
            else:
                genotype_dd_count += 1

    call_rate = float(total_genotype_count - missing_genotype_count) / \
        total_genotype_count

    if min_allele_id == '#N/A':
        fisher_hwe_p = '#N/A'
    else:
        fisher_hwe_p = fisher_hwe(
            total_genotype_count,
            genotype_Dd_count,
            min_allele_count,
            maj_allele_count,
        )

    return (call_rate, allele_num, min_allele_id, maj_allele_id,
            min_allele_freq, maj_allele_freq, fisher_hwe_p,
            genotype_DD_count, genotype_Dd_count,
            genotype_dd_count, missing_genotype_count,
            min_allele_count, maj_allele_count,
            missing_allele_count,)


def get_read_depth(info_field, vcf_format):
    read_depth_index = vcf_format.strip().split(':').index('DP')
    try:
        read_depth_field = info_field.strip().split(':')[read_depth_index]
    except:
        read_depth_field = None

    if not read_depth_field:
        genotype_index = vcf_format.strip().split(':').index('GT')
        genotype_field = info_field.strip().split(':')[genotype_index]

        for allele in genotype_field.split('/'):
            if allele != '.':
                raise ValueError('No read depth reported for genotyped sample.')

    return read_depth_field


def update_output_dict(output_dict, info_field, field_dict, group_name):

    info_fields = info_field.split(';')

    info_dict = dict()
    for field in info_fields:
        if len(field.split('=')) == 2:
            key, value = field.split('=')
            info_dict[key] = value

    for key, value in field_dict.items():
        if key in info_dict:
            output_dict.update({
                (group_name, value): info_dict[key],
            })


def read_gene_list(gene_list_fn):
    gene_set = set()

    with open(gene_list_fn) as f:
        for line in f:
            gene_set.add(line.strip())

    return gene_set


def read_vcf(input_fn, output_fn, gene_list_fn=None):

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Variants'

    if gene_list_fn:
        gene_set = read_gene_list(gene_list_fn)

    with open(input_fn) as f:
        x_row = 3

        for line in f:

            if line.startswith('##'):
                pass

            elif line.startswith('#'):

                vcf_headers = line[1:].strip().split('\t')
                samples = vcf_headers[9:]

                output_headers = []

                for group_name, fields in [
                    ('Position Info', POSITION_FIELDS),
                    ('Annotation', ANNOTATION_FIELDS),
                    ('Genotypes', samples),
                    ('Allele Frequency', ALLELE_FREQUENCY_FIELDS),
                    ('Read Depths', samples),
                    ('Impact', DBNSFP_IMPACT_FIELDS),
                    ('ExAc Frequency', EXAC_FREQUENCY_FIELDS),
                ]:
                    output_headers.append({
                        'group_name': group_name,
                        'fields': fields,
                    })

                write_headers(ws, output_headers)

            else:

                output_dict = dict()

                (
                    _chromosome,
                    _position,
                    _id,
                    _ref,
                    _alt,
                    _qual,
                    _filter,
                    _info,
                    _format,
                ) = line.strip().split('\t')[:9]
                sample_info = line.strip().split('\t')[9:]

                # QD not reported in all GATK lines
                # This is done when AD for all samples is 0
                try:
                    qd = _info.split('QD=')[1].split(';')[0]
                except:
                    qd = None

                _fields = (
                    ('Chromosome', _chromosome),
                    ('Position', _position),
                    ('dbSNP ID', _id),
                    ('Reference allele', _ref),
                    ('Alternative alleles', _alt),
                    ('QD',qd),
                    ('Filter', _filter),
                )
                for field, value in _fields:
                    output_dict.update({
                        ('Position Info', field): value,
                    })

                # Annotation fields, snpEff
                # Retrieve ANN fields
                _ann = _info.split('ANN=')[1].split(';')[0].split(',')
                _additional_reports = len(_ann) > 1

                entries = []
                for entry in _ann:
                    entry = entry.split('|')

                    distance = entry[14]
                    if not distance:
                        distance = '0'

                    entries.append({
                        'Allele': entry[0],
                        'Annotation': entry[1],
                        'Putative impact': entry[2],
                        'Gene Name': entry[3],
                        'Gene ID': entry[4],
                        'Feature Type': entry[5],
                        'Feature ID': entry[6],
                        'Transcript Biotype': entry[7],
                        'Exon or Intron Rank / Total': entry[8],
                        'HGVS DNA': entry[9],
                        'HGVS Protein': entry[10],
                        'cDNA Position / cDNA Length': entry[11],
                        'CDS Position / CDS Length': entry[12],
                        'Protein Position / Protein Length': entry[13],
                        'Distance to Feature': distance,
                    })

                translated_transcripts = \
                    [x for x in entries if all([
                        x['Transcript Biotype'] == 'protein_coding',
                        x['CDS Position / CDS Length'],
                        x['cDNA Position / cDNA Length'],
                    ])]

                if translated_transcripts:
                    entry = sorted(translated_transcripts, key=lambda x: (
                        int(x['Distance to Feature']),
                        SNPEFF_IMPACT_PRIORITY.index(x['Putative impact']),
                        -int(x['Transcript Biotype'] == 'protein_coding'),
                        -int(x['Feature Type'] == 'transcript'),
                        -int(x['CDS Position / CDS Length'].split('/')[1]),
                        -int(x['cDNA Position / cDNA Length'].split('/')[1]),
                        x['Feature ID'],
                    ))[0]
                else:
                    entry = sorted(entries, key=lambda x: (
                        int(x['Distance to Feature']),
                        SNPEFF_IMPACT_PRIORITY.index(x['Putative impact']),
                        -int(x['Transcript Biotype'] == 'protein_coding'),
                        -int(x['Feature Type'] == 'transcript'),
                        x['Feature ID'],
                    ))[0]

                for key, value in entry.items():
                    output_dict.update({
                        ('Annotation', key): value,
                    })
                output_dict.update({
                    ('Annotation', 'Additional Reports?'): _additional_reports,
                })

                # Genotypes
                genotypes = []
                for i, sample in enumerate(samples):
                    genotype = get_genotype(sample_info[i], _format, _ref, _alt)
                    output_dict.update({
                        #'{} Genotype'.format(sample): '_'.join(genotype),
                        ('Genotypes', sample): '_'.join(genotype),
                    })
                    genotypes.append(genotype)

                # Allele frequency
                (
                    call_rate,
                    allele_num,
                    min_allele_id,
                    maj_allele_id,
                    min_allele_freq,
                    maj_allele_freq,
                    fisher_hwe_p,
                    genotype_DD_count,
                    genotype_Dd_count,
                    genotype_dd_count,
                    missing_genotype_count,
                    min_allele_count,
                    maj_allele_count,
                    missing_allele_count,
                ) = get_allele_frequency_values(genotypes, _ref, _alt)

                _fields = (
                    ('Call Rate', call_rate),
                    ('Number of Distinct Alleles', allele_num),
                    ('Minor Allele', min_allele_id),
                    ('Major Allele', maj_allele_id),
                    ('Minor Allele Frequency', min_allele_freq),
                    ('Major Allele Frequency', maj_allele_freq),
                    ('Fisher\'s HWE P', fisher_hwe_p),
                    ('Genotype DD Count', genotype_DD_count),
                    ('Genotype Dd Count', genotype_Dd_count),
                    ('Genotype dd or Multi-Allelic Count', genotype_dd_count),
                    ('Missing Genotype Count', missing_genotype_count),
                    ('Minor Allele D Count', min_allele_count),
                    ('Major Allele d Count', maj_allele_count),
                    ('Missing Allele Count', missing_allele_count),
                )
                for field, value in _fields:
                    output_dict.update({
                        ('Allele Frequency', field): value,
                    })

                # Read depths
                for i, sample in enumerate(samples):
                    read_depth = get_read_depth(sample_info[i], _format)
                    output_dict.update({
                        ('Read Depths', sample): read_depth,
                    })

                # dbNSFP impact
                update_output_dict(output_dict, _info, DBNSFP_IMPACT_INFO_DICT, 'Impact')

                # ExAc frequencies
                update_output_dict(output_dict, _info, EXAC_FREQUENCY_INFO_DICT, 'ExAc Frequency')

                # FILTERS
                _filter = False

                # Gene set filter
                if gene_list_fn:
                    gene_name = output_dict[('Annotation', 'Gene Name')]
                    distance = output_dict[('Annotation', 'Distance to Feature')]
                    if distance != 0 or gene_name not in gene_set:
                        _filter = True

                if not _filter:
                    write_fields(ws, x_row, output_headers, output_dict, ref_allele=_ref)
                    x_row += 1

    ws.freeze_panes = ws['A3']
    wb.save(output_fn)


if __name__ == '__main__':

    parser = argparse.ArgumentParser()
    parser.add_argument('--gene_list', type=str, default=None, help='list of genes to keep')
    parser.add_argument('input_vcf_file', type=str, help='input vcf file')
    parser.add_argument('output_file', type=str, help='output file')
    args = parser.parse_args()

    read_vcf(
        args.input_vcf_file,
        args.output_file,
        gene_list_fn=args.gene_list,
    )
